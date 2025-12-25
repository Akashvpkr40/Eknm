import json
from openpyxl import load_workbook
from indic_transliteration import sanscript
from indic_transliteration.sanscript import transliterate

EXCEL_FILE = "Book List.xlsx"
OUTPUT_FILE = "data.js"

wb = load_workbook(EXCEL_FILE)
ws = wb.active

books = []

def prettify_itrans(text):
    text = text.replace("A", "aa").replace("I", "ee").replace("U", "oo")
    return text.capitalize()

for row in ws.iter_rows(min_row=2, values_only=True):
    title_ml = row[2]
    author = row[3]
    stock_no = row[1]
    price = row[6]

    if not title_ml or not stock_no:
        continue

    title_en_raw = transliterate(
        title_ml,
        sanscript.MALAYALAM,
        sanscript.ITRANS
    )

    title_en = prettify_itrans(title_en_raw)

    books.append({
        "name_ml": title_ml,
        "name_en": title_en,
        "author": author,
        "stockNumber": int(stock_no),
        "price": price,
        "image": f"books/{int(stock_no)}.jpg",
        "stock": "In Stock"
    })

with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
    f.write("const books = ")
    json.dump(books, f, ensure_ascii=False, indent=2)
    f.write(";")

print(f"✔ Exported {len(books)} bilingual book records to data.js")
